from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
from datetime import date


def generate_newsletter_excel(members_queryset):
    """Generate Excel export of active members for newsletter distribution"""

    # Split members into groups
    members_with_email = []
    members_without_email = []

    for member in members_queryset:
        if member.email and member.email.strip():
            members_with_email.append(member)
        else:
            members_without_email.append(member)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Column headers
    headers = [
        "MemberID",
        "FirstName",
        "LastName",
        "EmailName",
        "DateJoined",
        "Birthdate",
        "Expires",
        "MailName",
        "FullName",
    ]

    # Helper function to format date
    def format_date(d):
        if d:
            return d.strftime("%m/%d/%Y")
        return ""

    # Helper function to write member row
    def write_member_row(ws, member, row_num):
        mail_name = ""
        if member.email and member.email.strip():
            mail_name = f"{member.first_name} {member.last_name}<{member.email}>"

        ws.append(
            [
                member.member_id or "",
                member.first_name,
                member.last_name,
                member.email or "",
                format_date(member.date_joined),
                format_date(member.milestone_date),
                format_date(member.expiration_date),
                mail_name,
                f"{member.first_name} {member.last_name}",
            ]
        )

    # Create numbered sheets for members with emails (99 per sheet)
    if members_with_email:
        sheet_num = 1
        current_sheet = None
        row_count = 0

        for member in members_with_email:
            # Create new sheet if needed (every 99 rows)
            if row_count == 0:
                current_sheet = wb.create_sheet(title=f"Sheet {sheet_num}")
                # Write headers
                current_sheet.append(headers)
                # Make headers bold
                for cell in current_sheet[1]:
                    cell.font = Font(bold=True)
                row_count = 1

            # Write member row
            write_member_row(current_sheet, member, row_count + 1)
            row_count += 1

            # Start new sheet if we've reached 99 members
            if row_count > 99:
                sheet_num += 1
                row_count = 0

    # Create "No Email" sheet
    if members_without_email:
        no_email_sheet = wb.create_sheet(title="No Email")
        # Write headers
        no_email_sheet.append(headers)
        # Make headers bold
        for cell in no_email_sheet[1]:
            cell.font = Font(bold=True)

        # Write all members without emails
        for member in members_without_email:
            write_member_row(no_email_sheet, member, 0)

    # Ensure at least one sheet exists (for empty queryset case)
    if len(wb.sheetnames) == 0:
        empty_sheet = wb.create_sheet(title="Sheet 1")
        empty_sheet.append(headers)
        for cell in empty_sheet[1]:
            cell.font = Font(bold=True)

    # Save to BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create HTTP response
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="newsletter_data_{date.today().strftime("%Y_%m_%d")}.xlsx"'
    )

    return response
